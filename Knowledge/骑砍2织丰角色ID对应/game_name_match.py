import pandas as pd
import re
import difflib
import warnings
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

# 定义颜色填充
EXACT_MATCH_COLOR = "C6EFCE"  # 浅绿色 - 精确匹配
FUZZY_MATCH_COLOR = "FFEB9C"  # 浅黄色 - 模糊匹配
NO_MATCH_COLOR = "F2F2F2"     # 浅灰色 - 未匹配

def normalize_name(name):
    """标准化名字，移除空格和特殊字符"""
    if pd.isna(name):
        return ''
    # 移除空格
    name = str(name).replace(' ', '')
    # 移除可能的中英文标点
    name = re.sub(r'[。，、；：！？（）【】《》"\'.,;:!?()\[\]{}]', '', name)
    return name.strip()

def find_multiple_matches(query, candidates, threshold=0.5, n=5):
    """使用difflib找到多个匹配项，返回排序后的列表"""
    if not query or not candidates:
        return []
    
    # 使用difflib找到多个接近的匹配
    matches = difflib.get_close_matches(query, candidates, n=n, cutoff=threshold)
    
    # 计算每个匹配的相似度
    results = []
    for match in matches:
        seq = difflib.SequenceMatcher(None, query, match)
        similarity = seq.ratio() * 100  # 转换为百分比
        results.append((match, similarity))
    
    # 按相似度降序排序
    results.sort(key=lambda x: x[1], reverse=True)
    
    return results

def format_candidates(candidate_list, name_mapping, simple_to_original):
    """格式化候选列表为字符串"""
    if not candidate_list:
        return ""
    
    formatted = []
    for i, (candidate, score) in enumerate(candidate_list):
        # 获取原始简体名（非标准化）
        original_name = simple_to_original.get(candidate, candidate)
        # 获取繁体名
        traditional_name = name_mapping.get(candidate, "")
        
        if traditional_name:
            formatted.append(f"{i+1}. {original_name} → {traditional_name} ({score:.1f}%)")
        else:
            formatted.append(f"{i+1}. {original_name} ({score:.1f}%)")
    
    return "\n".join(formatted)

def main():
    # 1. 读取Excel文件
    excel_path = 'your_file.xlsx'  # 替换为您的文件路径
    
    # 读取两个sheet
    sheet1 = pd.read_excel(excel_path, sheet_name='shokuho')
    sheet2 = pd.read_excel(excel_path, sheet_name='taikou5')
    
    print(f"第一个Sheet行数: {len(sheet1)}")
    print(f"第二个Sheet行数: {len(sheet2)}")
    
    # 2. 准备数据
    # 第一个Sheet第4列（列索引3），移除空格
    sheet1_names = []
    sheet1_original_names = []  # 保留原始名字
    for i, row in sheet1.iterrows():
        name = row.iloc[3] if len(row) > 3 else ''
        if pd.notna(name):
            normalized = str(name).replace(' ', '')
            sheet1_names.append(normalized)
            sheet1_original_names.append(str(name))
        else:
            sheet1_names.append('')
            sheet1_original_names.append('')
    
    # 第二个Sheet，简体中文名（第3列，列索引2）
    sheet2_simplified_names = []
    sheet2_traditional_names = []
    sheet2_original_simplified = []  # 保留原始简体名
    for i, row in sheet2.iterrows():
        simplified = row.iloc[2] if len(row) > 2 else ''
        traditional = row.iloc[3] if len(row) > 3 else ''
        
        if pd.notna(simplified):
            sheet2_simplified_names.append(normalize_name(simplified))
            sheet2_traditional_names.append(traditional if pd.notna(traditional) else '')
            sheet2_original_simplified.append(str(simplified))
        else:
            sheet2_simplified_names.append('')
            sheet2_traditional_names.append('')
            sheet2_original_simplified.append('')
    
    # 3. 创建映射字典（简体->繁体）
    name_mapping = {}
    simple_to_original = {}  # 标准化名->原始简体名
    for simple, trad, original in zip(sheet2_simplified_names, sheet2_traditional_names, sheet2_original_simplified):
        if simple and trad:
            name_mapping[simple] = trad
            simple_to_original[simple] = original
    
    # 4. 添加历史人物手动映射（用于提高匹配准确率）
    historical_name_mapping = {
        # 简体->繁体 对应关系
        "木下秀吉": "豐臣秀吉",
        "羽柴秀吉": "豐臣秀吉",
        "藤吉郎": "豐臣秀吉",
        "浓姬": "歸蝶",
        "吉法师": "織田信長",  # 织田信长幼名
        "三法师": "織田秀信",  # 织田秀信
        "竹千代": "德川家康",  # 德川家康幼名
        # 可以根据您的数据添加更多
    }
    
    for simple, trad in historical_name_mapping.items():
        if simple not in name_mapping:
            name_mapping[simple] = trad
            simple_to_original[simple] = simple
    
    # 5. 匹配过程
    traditional_names = []
    match_types = []  # 匹配方式：精确/模糊/无匹配
    match_details = []  # 匹配详情
    match_scores = []  # 最佳相似度分数
    matched_simplified = []  # 匹配到的简体名
    multiple_candidates = []  # 多个候选匹配（按可能性排序）
    
    # 先尝试精确匹配
    for i, name1 in enumerate(sheet1_names):
        if name1 and name1 in name_mapping:
            traditional_names.append(name_mapping[name1])
            match_types.append("精确匹配")
            match_details.append(f"精确匹配: {simple_to_original.get(name1, name1)}")
            match_scores.append(100)
            matched_simplified.append(simple_to_original.get(name1, name1))
            multiple_candidates.append("")  # 精确匹配不需要候选
        else:
            traditional_names.append('')
            match_types.append("")
            match_details.append("")
            match_scores.append(0)
            matched_simplified.append("")
            multiple_candidates.append("")
    
    # 对于未匹配到的，尝试模糊匹配
    unmatched_indices = [i for i in range(len(sheet1_names)) 
                        if not match_types[i] and sheet1_names[i]]
    
    if unmatched_indices:
        print(f"\n开始模糊匹配，未匹配数量: {len(unmatched_indices)}")
        
        # 准备候选列表（移除已精确匹配的）
        candidate_names = [name for name in name_mapping.keys() 
                          if name not in [sheet1_names[i] for i in range(len(sheet1_names)) if match_types[i] == "精确匹配"]]
        
        for idx in unmatched_indices:
            name1 = sheet1_names[idx]
            
            # 获取多个候选匹配
            candidates = find_multiple_matches(name1, candidate_names, threshold=0.33, n=5)
            
            if candidates:
                # 使用最佳匹配作为主匹配
                best_match, best_score = candidates[0]
                
                traditional_names[idx] = name_mapping.get(best_match, "")
                match_types[idx] = "模糊匹配"
                match_scores[idx] = best_score
                matched_simplified[idx] = simple_to_original.get(best_match, best_match)
                
                # 格式化多个候选
                candidate_text = format_candidates(candidates, name_mapping, simple_to_original)
                multiple_candidates[idx] = candidate_text
                
                # 创建详细的匹配信息
                if traditional_names[idx]:
                    match_details[idx] = f"最佳匹配: {matched_simplified[idx]} → {traditional_names[idx]} ({best_score:.1f}%)"
                else:
                    match_details[idx] = f"最佳匹配: {matched_simplified[idx]} ({best_score:.1f}%)"
                
                # 从候选列表中移除最佳匹配，避免重复匹配
                if best_match in candidate_names:
                    candidate_names.remove(best_match)
                
                # 显示匹配信息
                if len(candidates) > 1:
                    print(f"模糊匹配: '{sheet1_original_names[idx]}' -> {len(candidates)}个候选")
                    for j, (candidate, score) in enumerate(candidates[:3]):
                        candidate_name = simple_to_original.get(candidate, candidate)
                        print(f"    候选{j+1}: {candidate_name} ({score:.1f}%)")
                else:
                    print(f"模糊匹配: '{sheet1_original_names[idx]}' -> '{matched_simplified[idx]}' ({best_score:.1f}%)")
            else:
                match_types[idx] = "无匹配"
                match_details[idx] = "未找到匹配项"
                multiple_candidates[idx] = "无匹配项"
    
    # 6. 将新列添加到第一个Sheet
    sheet1['繁体中文名'] = traditional_names
    sheet1['匹配方式'] = match_types
    sheet1['匹配详情'] = match_details
    sheet1['相似度'] = match_scores
    sheet1['匹配到的简体名'] = matched_simplified
    sheet1['多个候选匹配'] = multiple_candidates
    
    # 7. 统计匹配结果
    exact_count = match_types.count("精确匹配")
    fuzzy_count = match_types.count("模糊匹配")
    no_match_count = match_types.count("无匹配")
    
    print(f"\n匹配统计:")
    print(f"总角色数: {len(sheet1)}")
    print(f"精确匹配数: {exact_count}")
    print(f"模糊匹配数: {fuzzy_count}")
    print(f"无匹配数: {no_match_count}")
    print(f"总匹配率: {(exact_count + fuzzy_count)/len(sheet1)*100:.1f}%")
    
    # 统计模糊匹配的候选数量
    multi_candidate_counts = []
    for candidates in multiple_candidates:
        if candidates and "无匹配项" not in candidates:
            # 计算候选数量（每行是一个候选）
            count = len(candidates.split('\n'))
            multi_candidate_counts.append(count)
    
    if multi_candidate_counts:
        avg_candidates = sum(multi_candidate_counts) / len(multi_candidate_counts)
        print(f"平均每个模糊匹配候选数: {avg_candidates:.1f}")
    
    # 8. 保存结果到新文件，并设置颜色标记
    output_path = 'character_names_with_candidates.xlsx'
    
    # 创建Excel写入器
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 写入更新后的第一个sheet
        sheet1.to_excel(writer, sheet_name='character_names_table', index=False)
        # 写入原始的第二个sheet
        sheet2.to_excel(writer, sheet_name='Taikou5CharacterList', index=False)
        
        # 获取工作簿和工作表
        workbook = writer.book
        worksheet = writer.sheets['character_names_table']
        
        # 设置列宽
        column_widths = {
            'A': 15, 'B': 15, 'C': 15, 'D': 25,  # 原始列
            'E': 25,  # 繁体中文名
            'F': 15,  # 匹配方式
            'G': 40,  # 匹配详情
            'H': 10,  # 相似度
            'I': 25,  # 匹配到的简体名
            'J': 60,  # 多个候选匹配（设置较宽）
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # 为多个候选匹配列设置自动换行
        last_col_letter = get_column_letter(worksheet.max_column)  # 最后一列是"多个候选匹配"
        for row in range(2, len(sheet1) + 2):
            cell = worksheet.cell(row=row, column=worksheet.max_column)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 为匹配方式列设置颜色填充
        match_type_col = None
        for col_idx, cell in enumerate(worksheet[1], 1):  # 第一行是表头
            if cell.value == "匹配方式":
                match_type_col = col_idx
                break
        
        if match_type_col:
            for row_idx in range(2, len(sheet1) + 2):  # 从第2行开始（跳过表头）
                match_type = worksheet.cell(row=row_idx, column=match_type_col).value
                
                if match_type == "精确匹配":
                    fill = PatternFill(start_color=EXACT_MATCH_COLOR, end_color=EXACT_MATCH_COLOR, fill_type="solid")
                elif match_type == "模糊匹配":
                    fill = PatternFill(start_color=FUZZY_MATCH_COLOR, end_color=FUZZY_MATCH_COLOR, fill_type="solid")
                elif match_type == "无匹配":
                    fill = PatternFill(start_color=NO_MATCH_COLOR, end_color=NO_MATCH_COLOR, fill_type="solid")
                else:
                    fill = None
                
                # 为整行设置颜色
                if fill:
                    for col in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row_idx, column=col).fill = fill
        
        # 冻结首行
        worksheet.freeze_panes = 'A2'
    
    print(f"\n结果已保存到: {output_path}")
    
    # 9. 生成匹配报告
    report_path = '匹配报告_带候选项.txt'
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(f"角色名匹配报告（带候选项）\n")
        f.write("="*60 + "\n")
        f.write(f"总角色数: {len(sheet1)}\n")
        f.write(f"精确匹配数: {exact_count}\n")
        f.write(f"模糊匹配数: {fuzzy_count}\n")
        f.write(f"无匹配数: {no_match_count}\n")
        f.write(f"总匹配率: {(exact_count + fuzzy_count)/len(sheet1)*100:.1f}%\n")
        if multi_candidate_counts:
            f.write(f"平均每个模糊匹配候选数: {avg_candidates:.1f}\n")
        f.write("\n")
        
        f.write("模糊匹配详情（带候选项）:\n")
        f.write("="*60 + "\n")
        fuzzy_indices = [i for i, t in enumerate(match_types) if t == "模糊匹配"]
        
        for i in fuzzy_indices:
            original_name = sheet1_original_names[i]
            best_match = matched_simplified[i]
            best_score = match_scores[i]
            candidates = multiple_candidates[i]
            
            f.write(f"\n行{i+1}: '{original_name}'\n")
            f.write(f"  最佳匹配: {best_match} ({best_score:.1f}%)\n")
            f.write(f"  所有候选:\n")
            for line in candidates.split('\n'):
                f.write(f"    {line}\n")
        
        f.write("\n" + "="*60 + "\n")
        f.write("无匹配的条目:\n")
        f.write("-"*60 + "\n")
        no_match_indices = [i for i, t in enumerate(match_types) if t == "无匹配"]
        for i in no_match_indices[:50]:  # 只显示前50个
            original_name = sheet1_original_names[i]
            f.write(f"行{i+1}: {original_name}\n")
        
        if len(no_match_indices) > 50:
            f.write(f"  ... 还有{len(no_match_indices)-50}个未显示\n")
    
    print(f"详细匹配报告已保存到: {report_path}")
    
    # 10. 显示一些示例
    print("\n示例 - 模糊匹配带多个候选:")
    fuzzy_indices = [i for i, t in enumerate(match_types) if t == "模糊匹配"]
    if fuzzy_indices:
        for i in fuzzy_indices[:3]:  # 显示前3个示例
            original_name = sheet1_original_names[i]
            candidates = multiple_candidates[i]
            print(f"\n示例 {i+1}: '{original_name}'")
            for line in candidates.split('\n')[:3]:  # 只显示前3个候选
                print(f"  {line}")
    
    # 11. 建议的下一步操作
    print("\n建议:")
    print("1. 打开 Excel 文件，查看黄色标记的行（模糊匹配）")
    print("2. 检查 '多个候选匹配' 列，查看所有可能的匹配项")
    print("3. 候选项按相似度从高到低排序，第一个是最佳匹配")
    print("4. 您可以手动选择最合适的匹配，然后更新 '繁体中文名' 列")
    print("5. 对于太阁5中可能没有的角色，可以留空或标记")

if __name__ == "__main__":
    main()